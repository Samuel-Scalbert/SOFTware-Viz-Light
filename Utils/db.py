import os
import json
from tqdm import tqdm
import xml.etree.ElementTree as ET
import re
from openpyxl import load_workbook
from Utils.TEI_to_JSON import transformer_TEI_JSON
import requests


def get_publication_date(doi):
    url = f'https://api.crossref.org/works/{doi}'

    try:
        response = requests.get(url)

        if response.status_code == 200:
            data = response.json()
            publication_date = data['message'].get('issued', {}).get('date-parts', [])
            if publication_date:
                year = publication_date[0][0]
                return year
            else:
                return None
        else:
            return None

    except Exception as e:
        return f"Error: {str(e)}"

def check_or_create_collection(db, collection_name, collection_type='Collection'):
    """
    Checks if a collection exists in the database. If not, creates the collection.
    :param db: Database connection
    :param collection_name: Name of the collection
    :param collection_type: Type of the collection ('Collection' or 'Edges')
    :return: The collection object
    """
    if db.hasCollection(collection_name):
        return db[collection_name]
    else:
        db.createCollection(collection_type, name=collection_name)
        return db[collection_name]

def duplicates_JSON(lst):
    seen = set()
    duplicates = []

    for item in lst:
        item_hashable = str(item)
        if item_hashable in seen:
            duplicates.append(item)
        else:
            seen.add(item_hashable)

    return duplicates

def insert_json_db(data_path_json,data_path_xml,db):
    software_document = []
    list_errors = []
    #db.dropAllCollections()

    workbook = load_workbook(filename='./app/static/data/Logiciels_Blacklist_et_autres_remarques.xlsx')
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    blacklist = []
    for row in data[1:]:
        blacklist.append(row[0])

    # Create or retrieve collections
    documents_collection = check_or_create_collection(db, 'documents')
    softwares_collection = check_or_create_collection(db, 'softwares')
    references_collection = check_or_create_collection(db, 'references')
    authors_collection = check_or_create_collection(db, 'authors')
    structures_collection = check_or_create_collection(db, 'structures')

    # Create or retrieve edge collections
    doc_struc_edge = check_or_create_collection(db, 'edge_doc_to_struc', 'Edges')
    auth_struc_edge = check_or_create_collection(db, 'edge_auth_to_struc', 'Edges')
    doc_soft_edge = check_or_create_collection(db, 'edge_doc_to_software', 'Edges')
    doc_ref_edge = check_or_create_collection(db, 'edge_doc_to_reference', 'Edges')
    doc_auth_edge = check_or_create_collection(db, 'edge_doc_to_author', 'Edges')

    data_json_files = os.listdir(data_path_json)
    data_xml_list = os.listdir(data_path_xml)
    files_list_registered = db.AQLQuery('FOR hal_id in documents RETURN hal_id.file_hal_id', rawResults=True)
    dict_edge_author = {}

    for data_file_xml in tqdm(data_xml_list):
        file_path = f'{data_path_xml}/{data_file_xml}'
        file_name = os.path.basename(file_path)
        while "." in file_name:
            file_name, extension = os.path.splitext(file_name)

        if file_name in files_list_registered:
            pass
            #continue

        with (open(file_path, 'r', encoding='utf-8') as xml_file):
            data_json_get_document = {}
            tree = ET.parse(xml_file)
            root = tree.getroot()
            ns = {"tei": "http://www.tei-c.org/ns/1.0", 'xml': 'http://www.w3.org/XML/1998/namespace'}


            # DOCUMENT -----------------------------------------------------

            title = tree.find(".//tei:fileDesc//tei:titleStmt//tei:title", ns)
            title = title.text

            # Initialize the year variables
            final_year = None
            year = None
            year_sub = None
            year_doi = None
            year_pub = None

            # Regex pattern to extract a 4-digit year
            year_pattern = r'\b([12]\d{3})\b'

            # Check DOI for a date first (most reliable)
            doi_tag = tree.find(".//tei:fileDesc//tei:sourceDesc//tei:biblStruct//tei:idno[@type='DOI']", ns)
            if doi_tag is not None:
                year_doi = get_publication_date(doi_tag.text)
                if year_doi is not None:
                    final_year = year_doi
                    data_json_get_document['date'] = final_year
            else:
                # If no DOI, check for the submission date
                sub_tag = tree.find(".//tei:fileDesc//tei:sourceDesc//tei:biblStruct//tei:note[@type='submission']", ns)
                if sub_tag is not None:
                    match = re.search(year_pattern, sub_tag.text)
                    if match:
                        year_sub = match.group(1)
                        final_year = year_sub
                        data_json_get_document['date'] = final_year

                # If no DOI or submission date, check the publication date
                if final_year is None:
                    pub_tag = tree.find(".//tei:fileDesc//tei:publicationStmt//tei:date[@type='published']", ns)
                    if pub_tag is not None:
                        match = re.search(year_pattern, pub_tag.text)
                        if match:
                            if match.group(1) == pub_tag.attrib.get('when')[:4]:
                                year_pub = match.group(1)
                                final_year = year_pub
                                data_json_get_document['date'] = final_year

                # If no year from DOI, submission, or publication, check other sources
                if final_year is None:
                    date_tag = tree.findall(
                        ".//tei:fileDesc//tei:sourceDesc//tei:biblStruct//tei:monogr//tei:imprint//tei:date", ns)
                    for text in date_tag:
                        if text.text is not None:
                            match = re.search(year_pattern, text.text)
                            if match:
                                year = match.group(1)
                                final_year = year
                                data_json_get_document['date'] = final_year
                                break  # Stop once a valid date is found

            # ABSTRACT -----------------------------------------------------

            abstract = tree.find(".//{http://www.tei-c.org/ns/1.0}abstract")
            if abstract:
                tag_text = list(abstract)[0]
                if tag_text.tag == '{http://www.tei-c.org/ns/1.0}div':
                    for p_tag in list(tag_text):
                        text = "".join(p_tag.itertext())
                    data_json_get_document['abstract'] = ['GROBID' , text]

            data_json_get_document['file_hal_id'] = file_name
            data_json_get_document['title'] = title

            document_document = documents_collection.createDocument(data_json_get_document)
            document_document.save()

            # SOFTWARE -----------------------------------------------------

            if f"{file_name}.software.json" in data_json_files:
                with open(f'{data_path_json}/{file_name}.software.json', 'r') as json_file:
                    data_json = json.load(json_file)
                    data_json_get_mentions = data_json.get("mentions")

                    # Remove duplicates
                    for elm in duplicates_JSON(data_json_get_mentions):
                        data_json_get_mentions.remove(elm)

                    # Process each mention
                    for mention in data_json_get_mentions:
                        if mention['software-name']['normalizedForm'] not in blacklist:
                            mention['software_name'] = mention.pop('software-name')
                            mention['software_type'] = mention.pop('software-type')
                            software_document = softwares_collection.createDocument(mention)
                            software_document.save()

                            # Create edge from document to software
                            edge_doc_soft = doc_soft_edge.createEdge()
                            edge_doc_soft['_from'] = document_document._id
                            edge_doc_soft['_to'] = software_document._id
                            edge_doc_soft.save()

            # REFERENCES -----------------------------------------------------

                    # Process each reference
                    data_json_get_references = data_json.get("references")
                    for reference in data_json_get_references:
                        result_json = []
                        try:
                            result_json,error = transformer_TEI_JSON(reference['tei'])
                            if len(error) > 0:
                                list_errors.append(error, reference['tei'])
                        except Exception as e:
                            print(f"Error during the transformation from XML to JSON: {e}")
                        if result_json:
                            reference['json'] = result_json
                        references_document = references_collection.createDocument(reference)
                        references_document.save()

                        # Create edge from document to reference
                        edge_doc_ref = doc_ref_edge.createEdge()
                        edge_doc_ref['_from'] = document_document._id
                        edge_doc_ref['_to'] = references_document._id
                        edge_doc_ref.save()
                # Define the AQL query to fetch software names and their counts
                query = f"""
                FOR doc IN edge_software
                    FILTER doc._from == "{document_document._id}"
                    LET software = DOCUMENT(doc._to)
                    COLLECT softwareName = software.software_name.normalizedForm WITH COUNT INTO count
                    RETURN {{ softwareName, count }}
                """

            # AUTHORS -----------------------------------------------------

            author_list = tree.findall(".//tei:fileDesc//tei:sourceDesc//tei:biblStruct//tei:analytic//tei:author", ns)
            list_author_old = []

            for elm in author_list:
                author = {}

                # Extract author name
                author_name = {}
                persName = elm.find("{http://www.tei-c.org/ns/1.0}persName")
                if persName is not None:
                    for name in persName:
                        author_name[name.tag.split('}')[1]] = name.text

                # Construct a unique identifier for the author based on their name
                author_forename = f"{author_name.get('forename', '').strip()}"
                author_surname = f"{author_name.get('surname', '').strip()}"

                # Check if author is already registered
                result = db.AQLQuery(
                    f'FOR auth IN authors FILTER auth.name.surname == "{author_surname}" FILTER auth.name.forename == "{author_forename}" RETURN auth._id',
                    rawResults=True)

                if not result:
                    registered = False
                else:
                    registered = True

                # Extract author's role
                try:
                    role = elm.attrib['role']
                except KeyError:
                    role = 'unknown'

                # Document information
                document_halid = data_file_xml.replace(".hal.xml", "").replace(".hal.grobid.xml", "").replace(".xml",
                                                                                                              "").replace(
                    ".hal.xml", "")

                # Author document association
                author_documents = [{
                    'document_halid': document_halid,
                    'role': role
                }]

                # If not registered, create a new author document
                if registered:
                    author['name'] = author_name
                    author['documents'] = author_documents

                    document_author = authors_collection.createDocument(author)
                    document_author.save()

                    # Register the author by their name
                    author_document_id = document_author._id

                    # Create edge between document and author
                    edge_doc_auth = doc_auth_edge.createEdge()
                    edge_doc_auth['_from'] = document_document._id
                    edge_doc_auth['_to'] = document_author._id
                    edge_doc_auth.save()

                # If already registered, update their documents
                else:
                    # AQL query to append documents to the existing author
                    aql_query = f'''
                        FOR doc IN authors
                            FILTER doc._id == '{author_document_id}'
                            UPDATE doc WITH {{ 
                                documents: APPEND(doc.documents, {author_documents}, true)
                            }} IN authors
                    '''
                    # Execute the AQL query
                    result = db.AQLQuery(aql_query, rawResults=True)

                    # Create edge between document and author
                    edge_doc_auth = doc_auth_edge.createEdge()
                    edge_doc_auth['_from'] = document_document._id
                    edge_doc_auth['_to'] = author_document_id
                    edge_doc_auth.save()

                # STRUCT -----------------------------------------------------

                # Collect paths for all affiliations
                author_affiliation_paths = []
                affiliations_list = elm.findall("{http://www.tei-c.org/ns/1.0}affiliation")
                if affiliations_list:
                    for affiliation in affiliations_list:
                        list_org = affiliation.findall("{http://www.tei-c.org/ns/1.0}orgName")
                        for elm in list_org:
                            # ['department', 'institution', 'laboratory']
                            if elm.attrib['type'] in ['institution', 'laboratory']:
                                structure_name = elm.text
                                structure_type = elm.attrib['type']
                                result = db.AQLQuery(
                                    f'FOR struc IN structures FILTER struc.name == "{structure_name}" RETURN struc._id',
                                    rawResults=True)
                                if not result:
                                    # Add new structure if it doesn't exist
                                        org = {
                                            'name': structure_name,
                                            'type': structure_type,
                                        }
                                        document_org = structures_collection.createDocument(org)
                                        document_org.save()
                                        struct_id = document_org._id
                                else:
                                    struct_id = result[0]

                                    # Create edge between document and structure
                                query = f"""
                                            FOR edge IN edge_doc_to_struc
                                                FILTER edge._from == "{document_document._id}"
                                                RETURN edge._to
                                        """
                                list_registered_edge_struc_doc = db.AQLQuery(query, rawResults=True)

                                if struct_id not in list_registered_edge_struc_doc:
                                    edge_doc_struc = doc_struc_edge.createEdge()
                                    edge_doc_struc['_from'] = document_document._id
                                    edge_doc_struc['_to'] = struct_id
                                    edge_doc_struc.save()
                                    # Create edge between author and structure
                                    query = f"""
                                                                   FOR edge IN edge_auth_to_struc
                                                                       FILTER edge._from == "{author_document_id}"
                                                                       RETURN edge._to
                                                               """
                                    list_registered_edge_struc_author = db.AQLQuery(query, rawResults=True)

                                    if struct_id not in list_registered_edge_struc_author:
                                        edge_auth_struc = auth_struc_edge.createEdge()
                                        edge_auth_struc['_from'] = author_document_id
                                        edge_auth_struc['_to'] = struct_id
                                        edge_auth_struc.save()

    if len(list_errors) > 0:
       print(list_errors)
